import streamlit as st
import pdfplumber
import pandas as pd
import re
import io

# ============================================================
# Constants
# ============================================================

CATEGORIES = {"Bunga", "Pajak", "Transfer", "Pembayaran", "Payment"}

PAGE_HEADER_KEYWORDS = [
    "REKENING KORAN",
    "S/N S01-",
    "halaman",
    "Ketentuan Umum",
    "Syarat Layanan Perbankan",
    "PT Bank Seabank",
    "Lembaga Penjamin Simpanan",
    "Otoritas Jasa Keuangan",
    "https://www.seabank.co.id",
]

# Regex patterns
DATE_PATTERN = re.compile(
    r"^(\d{2}\s+(?:JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC))\s", re.I
)
# Cocokkan angka format Indonesia (1.234.567) DAN angka kecil biasa (999, 684, dll)
AMOUNT_PATTERN = re.compile(r"\b((?:\d{1,3}\.)*\d{1,3})\b")
FULL_DATE_HEADER = re.compile(r"^\d{2}\s+\w{3}\s+\d{4}$")  # e.g. "01 APR 2026"
ACCOUNT_NUMBER = re.compile(r"^\d{10,}$")
COLUMN_HEADER = re.compile(r"^TANGGAL\s+TRANSAKSI")


# ============================================================
# Parser Functions
# ============================================================


def parse_statement(file_obj) -> tuple[dict, list[dict]]:
    """
    Parse Seabank e-statement PDF.
    Returns (account_info, transactions).
    """
    with pdfplumber.open(file_obj) as pdf:
        all_text = ""
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                all_text += text + "\n"

    lines = [line.strip() for line in all_text.split("\n")]

    account_info = _parse_account_info(lines)
    summary = _parse_summary(lines)
    account_info.update(summary)

    transactions = _parse_transactions(lines, account_info["saldo_awal"])

    return account_info, transactions


def _parse_account_info(lines: list[str]) -> dict:
    """Extract account holder info dari PDF header."""
    info = {
        "nama": "",
        "no_rekening": "",
        "alamat": "",
    }

    for i, line in enumerate(lines):
        if not info["nama"] and line and not any(
            kw in line for kw in PAGE_HEADER_KEYWORDS
        ):
            if not re.match(r"^(REKENING|S/N|01 APR|\d{2} \w{3})", line):
                info["nama"] = line
                break

    for line in lines:
        m = re.search(r"NO\.\s*REKENING\s*SEABANK:\s*(\d+)", line)
        if m:
            info["no_rekening"] = m.group(1)
            break

    if info["nama"]:
        in_address = False
        address_parts = []
        for line in lines:
            if line == info["nama"]:
                in_address = True
                continue
            if in_address:
                if "NO. REKENING" in line or "RINGKASAN" in line:
                    break
                if "Hubungi kami" in line:
                    parts = line.split("Hubungi kami")[0].strip()
                    if parts:
                        address_parts.append(parts)
                    continue
                if "live chat" in line:
                    break
                if line:
                    address_parts.append(line)

        info["alamat"] = ", ".join(address_parts) if address_parts else ""

    return info


def _parse_summary(lines: list[str]) -> dict:
    """Parse section RINGKASAN REKENING."""
    summary = {
        "periode": "",
        "saldo_awal": 0,
        "total_keluar": 0,
        "total_masuk": 0,
        "saldo_akhir": 0,
    }

    for line in lines:
        if "sampai" in line.lower():
            summary["periode"] = line.strip()

        if "TABUNGAN" in line:
            amounts = re.findall(r"(\d{1,3}(?:\.\d{3})+)", line)
            if len(amounts) == 4:
                summary["saldo_awal"] = int(amounts[0].replace(".", ""))
                summary["total_keluar"] = int(amounts[1].replace(".", ""))
                summary["total_masuk"] = int(amounts[2].replace(".", ""))
                summary["saldo_akhir"] = int(amounts[3].replace(".", ""))

    return summary


def _parse_transactions(lines: list[str], saldo_awal: int) -> list[dict]:
    """
    Parse semua transaksi dari section 'TABUNGAN - RINCIAN TRANSAKSI'.
    Menggunakan perbandingan saldo untuk menentukan KELUAR vs MASUK.
    """
    in_transactions = False
    tx_lines = []

    for line in lines:
        if "TABUNGAN - RINCIAN TRANSAKSI" in line:
            in_transactions = True
            continue
        if "Ketentuan Umum" in line:
            break
        if in_transactions:
            if COLUMN_HEADER.search(line):
                continue
            tx_lines.append(line)

    transactions = []
    desc_buffer: list[str] = []
    prev_saldo = saldo_awal

    for line in tx_lines:
        line = line.strip()
        if not line:
            continue

        if any(kw in line for kw in ["REKENING KORAN", "halaman"]):
            continue
        if line.startswith("S/N "):
            continue
        if FULL_DATE_HEADER.match(line):
            continue
        if "Hubungi kami" in line or "live chat" in line:
            continue

        m = DATE_PATTERN.match(line)
        if m:
            date_str = m.group(1).upper()
            # Extract amounts dari sisa baris (setelah tanggal)
            rest = line[m.end():]
            amounts = AMOUNT_PATTERN.findall(rest)

            if len(amounts) >= 2:
                amount_val = int(amounts[0].replace(".", ""))
                saldo_val = int(amounts[-1].replace(".", ""))

                # Tentukan KELUAR vs MASUK dengan membandingkan saldo sebelumnya
                if saldo_val > prev_saldo:
                    masuk = amount_val
                    keluar = 0
                elif saldo_val < prev_saldo:
                    keluar = amount_val
                    masuk = 0
                else:
                    keluar = 0
                    masuk = 0

                prev_saldo = saldo_val

                # Extract deskripsi tambahan dari baris tanggal
                extra_desc = rest
                for amt in amounts:
                    extra_desc = extra_desc.replace(amt, "", 1)
                extra_desc = re.sub(r"\d{10,}", "", extra_desc)
                extra_desc = extra_desc.strip()

                if extra_desc and extra_desc not in CATEGORIES:
                    desc_buffer.append(extra_desc)

                desc = " - ".join(desc_buffer) if desc_buffer else "-"
                transactions.append(
                    {
                        "TANGGAL": date_str,
                        "TRANSAKSI": desc,
                        "KELUAR (IDR)": keluar,
                        "MASUK (IDR)": masuk,
                        "SALDO AKHIR (IDR)": saldo_val,
                    }
                )
                desc_buffer = []

        elif line in CATEGORIES or ACCOUNT_NUMBER.match(line):
            continue
        else:
            desc_buffer.append(line)

    return transactions


# ============================================================
# Formatter Functions
# ============================================================


def format_idr(value: int) -> str:
    """Format integer as Indonesian Rupiah string."""
    if value == 0:
        return "-"
    return f"{value:,.0f}".replace(",", ".")


def format_idr_csv(value: int) -> str:
    """Format integer untuk CSV (tanpa titik)."""
    if value == 0:
        return "0"
    return str(value)


# ============================================================
# Streamlit UI
# ============================================================


def mainSeaBankEstatement():
    st.markdown(
        """
        <style>
            .metric-card {
                background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%);
                border-radius: 12px;
                padding: 20px;
                text-align: center;
                border: 1px solid #dee2e6;
            }
            .metric-label {
                font-size: 0.85rem;
                color: #6c757d;
                margin-bottom: 5px;
                font-weight: 500;
            }
            .metric-value {
                font-size: 1.25rem;
                font-weight: 700;
                color: #212529;
            }
        </style>
        """,
        unsafe_allow_html=True,
    )

    st.title("🏦 Seabank E-Statement Parser")
    st.caption("Upload e-statement PDF SeaBank untuk mengekstrak data transaksi")

    # --- File Upload ---
    uploaded_file = st.file_uploader(
        "Upload file PDF e-statement",
        type=["pdf"],
        help="Pilih file PDF e-statement dari SeaBank",
        label_visibility="collapsed",
    )

    if not uploaded_file:
        st.info("📋 Silakan upload file PDF e-statement SeaBank untuk memulai.")
        st.stop()

    # --- Parse ---
    with st.spinner("Membaca dan memproses e-statement..."):
        try:
            account_info, transactions = parse_statement(uploaded_file)
        except Exception as e:
            st.error(f"❌ Gagal membaca file PDF: {e}")
            st.stop()

    if not transactions:
        st.warning("⚠️ Tidak ditemukan transaksi dalam file ini.")
        st.stop()

    # --- Account Info ---
    st.markdown("---")
    st.subheader("📋 Informasi Rekening")

    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown(f"**Nama:** {account_info.get('nama', '-')}")
    with col2:
        st.markdown(f"**No. Rekening:** `{account_info.get('no_rekening', '-')}`")
    with col3:
        st.markdown(f"**Periode:** {account_info.get('periode', '-')}")

    if account_info.get("alamat"):
        st.markdown(f"**Alamat:** {account_info['alamat']}")

    # --- Summary Cards ---
    st.markdown("---")
    st.subheader("📊 Ringkasan Rekening")

    col_s1, col_s2, col_s3, col_s4 = st.columns(4)

    with col_s1:
        st.markdown(
            f"""
            <div class="metric-card">
                <div class="metric-label">Saldo Awal (IDR)</div>
                <div class="metric-value">{format_idr(account_info['saldo_awal'])}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    with col_s2:
        st.markdown(
            f"""
            <div class="metric-card">
                <div class="metric-label">Total Keluar (IDR)</div>
                <div class="metric-value" style="color: #dc3545;">
                    ({format_idr(account_info['total_keluar'])})
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    with col_s3:
        st.markdown(
            f"""
            <div class="metric-card">
                <div class="metric-label">Total Masuk (IDR)</div>
                <div class="metric-value" style="color: #198754;">
                    {format_idr(account_info['total_masuk'])}
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    with col_s4:
        st.markdown(
            f"""
            <div class="metric-card">
                <div class="metric-label">Saldo Akhir (IDR)</div>
                <div class="metric-value" style="color: #0d6efd;">
                    {format_idr(account_info['saldo_akhir'])}
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    # --- DataFrame ---
    df = pd.DataFrame(transactions)

    # --- Filter Controls ---
    st.markdown("---")
    st.subheader("🔍 Rincian Transaksi")

    filter_col1, filter_col2, filter_col3 = st.columns([2, 2, 1])

    with filter_col1:
        search_term = st.text_input(
            "Cari transaksi",
            placeholder="Ketik kata kunci (misal: Bunga, Transfer, Shopee...)",
            label_visibility="collapsed",
        )

    with filter_col2:
        unique_dates = sorted(df["TANGGAL"].unique())
        selected_date = st.selectbox(
            "Filter tanggal",
            options=["Semua Tanggal"] + unique_dates,
            label_visibility="collapsed",
        )

    with filter_col3:
        tx_type = st.selectbox(
            "Tipe transaksi",
            options=["Semua", "Keluar saja", "Masuk saja"],
            label_visibility="collapsed",
        )

    # Apply filters
    filtered_df = df.copy()

    if search_term:
        filtered_df = filtered_df[
            filtered_df["TRANSAKSI"].str.contains(search_term, case=False, na=False)
        ]

    if selected_date != "Semua Tanggal":
        filtered_df = filtered_df[filtered_df["TANGGAL"] == selected_date]

    if tx_type == "Keluar saja":
        filtered_df = filtered_df[filtered_df["KELUAR (IDR)"] > 0]
    elif tx_type == "Masuk saja":
        filtered_df = filtered_df[filtered_df["MASUK (IDR)"] > 0]

    # --- Display Table ---
    display_df = filtered_df.copy()
    display_df["KELUAR (IDR)"] = display_df["KELUAR (IDR)"].apply(format_idr)
    display_df["MASUK (IDR)"] = display_df["MASUK (IDR)"].apply(format_idr)
    display_df["SALDO AKHIR (IDR)"] = display_df["SALDO AKHIR (IDR)"].apply(format_idr)

    column_config = {
        "TANGGAL": st.column_config.TextColumn("TANGGAL", width="medium"),
        "TRANSAKSI": st.column_config.TextColumn("TRANSAKSI", width="large"),
        "KELUAR (IDR)": st.column_config.TextColumn("KELUAR (IDR)", width="large"),
        "MASUK (IDR)": st.column_config.TextColumn("MASUK (IDR)", width="large"),
        "SALDO AKHIR (IDR)": st.column_config.TextColumn("SALDO AKHIR (IDR)", width="large"),
    }

    st.dataframe(
        display_df,
        use_container_width=True,
        hide_index=True,
        column_config=column_config,
        height=500,
    )

    st.caption(f"Menampilkan {len(filtered_df)} dari {len(df)} transaksi")

    # --- Filtered Summary ---
    if len(filtered_df) != len(df):
        st.markdown("---")
        f_col1, f_col2 = st.columns(2)
        with f_col1:
            total_keluar_filtered = filtered_df["KELUAR (IDR)"].sum()
            st.metric("Total Keluar (filter)", format_idr(total_keluar_filtered))
        with f_col2:
            total_masuk_filtered = filtered_df["MASUK (IDR)"].sum()
            st.metric("Total Masuk (filter)", format_idr(total_masuk_filtered))

    # --- Export Buttons ---
    st.markdown("---")
    st.subheader("💾 Export Data")

    export_col1, export_col2, export_col3 = st.columns(3)

    with export_col1:
        csv_data = _to_csv(df)
        st.download_button(
            label="📥 Export CSV (Semua Data)",
            data=csv_data,
            file_name=f"seabank_statement_{_safe_filename(account_info)}.csv",
            mime="text/csv",
            use_container_width=True,
        )

    with export_col2:
        if not filtered_df.empty:
            csv_filtered = _to_csv(filtered_df)
            st.download_button(
                label="📥 Export CSV (Data Terfilter)",
                data=csv_filtered,
                file_name=f"seabank_statement_filtered_{_safe_filename(account_info)}.csv",
                mime="text/csv",
                use_container_width=True,
            )

    with export_col3:
        try:
            excel_data = _to_excel(df)
            st.download_button(
                label="📊 Export Excel (Semua Data)",
                data=excel_data,
                file_name=f"seabank_statement_{_safe_filename(account_info)}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        except ImportError:
            st.warning("Install openpyxl untuk export Excel: pip install openpyxl")

    # --- Raw Data ---
    with st.expander("🔧 Data Mentah (Raw JSON)"):
        st.json(transactions)


# ============================================================
# Export Helpers
# ============================================================


def _to_csv(df: pd.DataFrame) -> str:
    """Convert DataFrame ke CSV string."""
    export_df = df.copy()
    export_df["KELUAR (IDR)"] = export_df["KELUAR (IDR)"].apply(format_idr_csv)
    export_df["MASUK (IDR)"] = export_df["MASUK (IDR)"].apply(format_idr_csv)
    export_df["SALDO AKHIR (IDR)"] = export_df["SALDO AKHIR (IDR)"].apply(format_idr_csv)
    return export_df.to_csv(index=False)


def _to_excel(df: pd.DataFrame) -> bytes:
    """Convert DataFrame ke Excel bytes dengan formatting."""
    export_df = df.copy()
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="Transaksi")

        ws = writer.sheets["Transaksi"]

        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        # Header styling
        header_fill = PatternFill(
            start_color="1F4E79", end_color="1F4E79", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col in range(1, 6):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Column widths
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 22
        ws.column_dimensions["E"].width = 25

        number_format = "#,##0"
        even_fill = PatternFill(
            start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"
        )
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        for row in range(2, len(export_df) + 2):
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border

                if col in (3, 4, 5):
                    cell.number_format = number_format
                    cell.alignment = Alignment(horizontal="right")
                elif col == 1:
                    cell.alignment = Alignment(horizontal="center")

            if row % 2 == 0:
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = even_fill

    output.seek(0)
    return output.read()


def _safe_filename(account_info: dict) -> str:
    """Generate safe filename dari account info."""
    no_rek = account_info.get("no_rekening", "unknown")
    periode = account_info.get("periode", "")
    dates = re.findall(r"\d{2}\s+\w{3}\s+\d{4}", periode)
    if dates:
        parts = []
        for d in dates:
            parts.append(re.sub(r"\s+", "_", d.lower()))
        return f"{'_'.join(parts)}_{no_rek}"
    return no_rek


# ============================================================
# Entry Point
# ============================================================

if __name__ == "__main__":
    mainSeaBankEstatement()
