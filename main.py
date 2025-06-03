from tabula import read_pdf
import pandas as pd
import numpy as np
import os
from tqdm import tqdm
from openpyxl import load_workbook
import streamlit as st


def is_currency(value):
    if pd.isna(value) or value == '':
        return False
    try:
        float(str(value).replace(',', ''))
    except ValueError:
        return False
    else:
        return True


def clean_numeric_columns(dataframe, columns):

    for column in columns:
        dataframe[column] = dataframe[column].str.replace(',', '')
        dataframe[column] = pd.to_numeric(dataframe[column], errors='coerce')
        dataframe[column] = dataframe[column].astype('float')

    return dataframe


def union_source(dataframes):

    dfs = []
    for temp_df in dataframes:

        # Split DB into new column
        temp_df[['amount', 'type']] = temp_df[4].str.extract(r'([\d,]+(?:\.\d+)?)\s*(DB|CR)?')
        temp_df = temp_df.drop(temp_df.columns[4], axis=1)

        if(len(temp_df.columns) == 7):
            # Name column and reorder
            temp_df.columns = ['date', 'desc', 'detail', 'branch', 'balance', 'amount', 'type']
            temp_df = temp_df[['date', 'desc', 'detail', 'branch', 'amount', 'type', 'balance']]

            dfs.append(temp_df)

    df = pd.concat(dfs, ignore_index=True)
    df = df.fillna(value=np.nan)
                
    return df


def insert_shifted_column(dataframe):

    # Add new columns with shifted values for comparison
    dataframe['prev_date'] = dataframe['date'].shift(1)
    dataframe['prev_desc'] = dataframe['desc'].shift(1)
    dataframe['prev_detail'] = dataframe['detail'].shift(1)
    dataframe['prev_branch'] = dataframe['branch'].shift(1)
    dataframe['prev_amount'] = dataframe['amount'].shift(1)
    dataframe['prev_transaction_type'] = dataframe['type'].shift(1)
    dataframe['prev_balance'] = dataframe['balance'].shift(1)

    dataframe = dataframe.fillna(value=np.nan)

    return dataframe


def extract_transactions(dataframe):

    transactions = []
    details = []
    descs = []
    temp = {}
    last_transaction = None  # Variable to track the last added transaction

    for index, row in dataframe.iterrows():

        if row['desc'] == 'S':
            transaction = {
                "date": temp['date'],
                "desc": ' | '.join(descs).strip(' | ') if descs else '',
                "detail": ' | '.join(details).strip(' | ') if details else '',
                "branch": temp['branch'],
                "amount": temp['amount'],
                "transaction_type": temp['transaction_type'] if temp['transaction_type'] == 'DB' else 'CR',
                "balance": temp['balance']
            }
            if transaction != last_transaction:  # Avoid duplicate addition
                transactions.append(transaction)
                last_transaction = transaction
            break

        if row['desc'] == 'SALDO AWAL':
            transaction = {
                "date": row['date'],
                "desc": 'SALDO AWAL',
                "detail": '',
                "branch": '',
                "amount": '',
                "transaction_type": '',
                "balance": row['balance']
            }
            transactions.append(transaction)
            continue

        if row['desc'] == 'KETE':
            continue
        
        # New Transaction
        if not pd.isna(row['amount']) and (
            pd.isna(row['prev_amount']) or
            row['amount'] != row['prev_amount'] or
            row['desc'] != row['prev_desc'] or
            row['detail'] != row['prev_detail'] or
            row['branch'] != row['prev_branch']
        ):
            # Save previous transaction
            if temp:
                transaction = {
                    "date": temp['date'],
                    "desc": ' | '.join(descs).strip(' | ') if descs else '',
                    "detail": ' | '.join(details).strip(' | ') if details else '',
                    "branch": temp['branch'],
                    "amount": temp['amount'],
                    "transaction_type": temp['transaction_type'] if temp['transaction_type'] == 'DB' else 'CR',
                    "balance": temp['balance']
                }
                if transaction != last_transaction:  # Avoid duplicate addition
                    transactions.append(transaction)
                    last_transaction = transaction
                details = []
                descs = []
                temp = {}

            temp = {
                'date': row['date'],
                'branch': row['branch'],
                'amount': row['amount'],
                'transaction_type': row['type'],
                'balance': row['balance']
            }

        if not pd.isna(row['desc']) and row['desc'].strip():
            descs.append(row['desc'])

        if not pd.isna(row['detail']) and row['detail'].strip():
            details.append(row['detail'])

    # Save the last transaction if it hasn't been added yet
    if temp:
        transaction = {
            "date": temp['date'],
            "desc": ' | '.join(descs).strip(' | ') if descs else '',
            "detail": ' | '.join(details).strip(' | ') if details else '',
            "branch": temp['branch'],
            "amount": temp['amount'],
            "transaction_type": temp['transaction_type'] if temp['transaction_type'] == 'DB' else 'CR',
            "balance": temp['balance']
        }
        if transaction != last_transaction:  # Avoid duplicate addition
            transactions.append(transaction)

    transaction_dataframe = pd.DataFrame(transactions)

    return transaction_dataframe


def calculate_balance(dataframe, init_balance):
    dataframe['balance'] = init_balance
    # Iterate over rows
    for index, row in dataframe.iterrows():
        # If transaction type is 'DB', subtract amount from balance
        if row['transaction_type'] == 'DB':
            if index == 0:
                # For the first row, subtract amount from init_balance
                dataframe.at[index, 'balance'] -= row['amount']
            else:
                # For subsequent rows, subtract amount from the previous row's balance
                dataframe.at[index, 'balance'] = dataframe.at[index - 1, 'balance'] - row['amount']
        # If transaction type is 'CR', add amount to balance
        elif row['transaction_type'] == 'CR':
            if index == 0:
                # For the first row, add amount to init_balance
                dataframe.at[index, 'balance'] += row['amount']
            else:
                # For subsequent rows, add amount to the previous row's balance
                dataframe.at[index, 'balance'] = dataframe.at[index - 1, 'balance'] + row['amount']

    return dataframe


def save_to_excel(dataframe, output_filename, sheet_name):
    if os.path.isfile(output_filename):
        writer = pd.ExcelWriter(output_filename, engine="openpyxl", mode='a', if_sheet_exists='replace')
    else:
        writer = pd.ExcelWriter(output_filename, engine="openpyxl")

    dataframe.to_excel(writer, sheet_name=sheet_name, index=False)

    workbook = writer.book
    worksheet = writer.sheets[sheet_name]

    # Format column into currency IDR
    for cell in worksheet['E']:
        cell.number_format = '_-Rp* #,##0.00_-;[Red]-Rp* #,##0.00_-;_-Rp* "-"_-;_-@_-'
    for cell in worksheet['G']:
        cell.number_format = '_-Rp* #,##0.00_-;[Red]-Rp* #,##0.00_-;_-Rp* "-"_-;_-@_-'

    # Autofit column width
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value)) for cell in column_cells)
        if str(column_cells[0].column) in ['5', '7']:
            worksheet.column_dimensions[column_cells[0].column_letter].width = max_length + 10
        else:
            worksheet.column_dimensions[column_cells[0].column_letter].width = max_length + 2

    writer.close()

    return


def get_year_month(sheet_name):
    parts = sheet_name.split(' ', 1)  # Split into at most two parts
    if len(parts) != 2:
        raise ValueError(f"Invalid sheet name format: {sheet_name}")
    year, month_name = parts
    month_dict = {
        'JANUARI': 1, 'FEBRUARI': 2, 'MARET': 3, 'APRIL': 4,
        'MEI': 5, 'JUNI': 6, 'JULI': 7, 'AGUSTUS': 8,
        'SEPTEMBER': 9, 'OKTOBER': 10, 'NOVEMBER': 11, 'DESEMBER': 12
    }
    if month_name not in month_dict:
        raise ValueError(f"Invalid month name in sheet name: {month_name}")
    return int(year), month_dict[month_name]


def reorder_sheets(output_filename):

    wb = load_workbook(output_filename)
    sheet_names = wb.sheetnames
    sorted_sheets = sorted(sheet_names, key=get_year_month, reverse=True)
    wb._sheets.sort(key=lambda x: sorted_sheets.index(x.title))
    wb.save(output_filename)

    return

statements_folder = "statements"

def main():
    st.title("BCA e-Statement Converter")

    # File uploader
    uploaded_files = st.file_uploader("Upload PDF Statements", type="pdf", accept_multiple_files=True)

    if uploaded_files:
        all_transactions = []

        for uploaded_file in uploaded_files:
            file_path = f"temp_{uploaded_file.name}"
            with open(file_path, "wb") as f:
                f.write(uploaded_file.read())

            # Process the uploaded file
            header_dataframe = read_pdf(file_path, area=(70, 315, 141, 548), pages='1', pandas_options={'header': None, 'dtype': str}, force_subprocess=True)[0]
            periode = header_dataframe.loc[header_dataframe[0] == 'PERIODE', 2].values[0]
            periode = ' '.join(reversed(periode.split()))
            no_rekening = header_dataframe.loc[header_dataframe[0] == 'NO. REKENING', 2].values[0]

            dataframes = read_pdf(file_path, area=(231, 25, 797, 577), columns=[86, 184, 300, 340, 467], pages='all', pandas_options={'header': None, 'dtype': str}, force_subprocess=True)

            init_balance = dataframes[0].loc[dataframes[0][1] == 'SALDO AWAL', 5].values[0]
            init_balance = float(init_balance.replace(',', ''))

            df = union_source(dataframes)
            df = clean_numeric_columns(df, ['amount', 'balance'])
            df = insert_shifted_column(df)

            transaction_dataframe = extract_transactions(df)
            transaction_dataframe = transaction_dataframe.drop('balance', axis=1)
            transaction_dataframe = calculate_balance(transaction_dataframe, init_balance)

            transaction_dataframe['source_file'] = uploaded_file.name
            all_transactions.append(transaction_dataframe)

        # Combine all transactions
        global_dataframe = pd.concat(all_transactions, ignore_index=True)

        # Display the dataframe
        st.write("### Combined Transactions")
        st.dataframe(global_dataframe)

        # Download button
        output_filename = "Combined_Statements.xlsx"
        with pd.ExcelWriter(output_filename, engine="openpyxl") as writer:
            global_dataframe.to_excel(writer, sheet_name="All Transactions", index=False)

        with open(output_filename, "rb") as f:
            st.download_button(
                label="Download Excel File",
                data=f,
                file_name=output_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

if __name__ == "__main__":
    main()